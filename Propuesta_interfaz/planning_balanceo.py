"""
planning_balanceo.py
Balanceo de demanda de jeringas entre las lineas DC1 y DC2 (sitio Naucalpan).

Como funciona, en cuatro pasos:
    1. Del forecast se toman las columnas de meses futuros (el bloque de piezas).
    2. Cada SKU se cruza con el catalogo para saber su familia de granel y si
       corre en DC2 (ambas banderas de linea encendidas) o solo en DC1.
    3. La demanda se suma por familia y linea. DC2 se deja en lotes completos y
       el remanente en piezas se pasa al bloque DC1 de la misma familia.
    4. Todo se divide entre el tamano de lote y se exporta a un Excel.

Requisitos: pip install pandas openpyxl
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Parametros de negocio (la app puede pasar otros) ---
TAMANO_LOTE = 115_200      # jeringas por lote
FORMATO_MES = "%m-%Y"

# --- Nombres de columna ---
id_cols = ['SKUMERCK','Description']

# --- Formato del Excel de salida ---
FUENTE = "Arial"
VERDE_MERCK = "0F5A3C"
FORMATO_LOTES = "#,##0.00"
FORMATO_PIEZAS = "#,##0"

#Joe:la funcino log solo sirve para notificar al usuario, no es necesaria
def log(mensaje):
    """Unico punto de salida de mensajes. La app de escritorio reemplaza esta
    funcion por la suya para mostrarlos en la bitacora de la ventana."""
    print(mensaje)


# ==============================================================================
# 1. Lectura
# ==============================================================================

def leer_entradas(ruta_fcst, hoja_fcst, ruta_catalogo, hoja_catalogo):
    """Lee los dos Excel y revisa que traigan lo minimo necesario."""
    fcst = pd.read_excel(ruta_fcst, sheet_name=hoja_fcst)
    fcst['SKUMERCK']=fcst['SKUMERCK'].astype(str).str.strip()
    catalogo = pd.read_excel(ruta_catalogo, sheet_name=hoja_catalogo)
    catalogo['SKUMERCK']=catalogo['SKUMERCK'].astype(str).str.strip()

    #Joe: Esta seccion checa si hay algun faltante en el catalogo
    #Joe: no es necesaria pero ayuda, parte de la alucinacion de la ia
    faltantes = [c for c in ('SKUMERCK', "Nombre granel", "Linea Granel 1", "Linea Granel 2")
                 if c not in catalogo.columns]
    if faltantes:
        raise ValueError(f"Al catalogo le faltan columnas: {faltantes}")
    #Joe: busca duplicados en el catalogo, no es necesaria pero ayuda,
    #Joe:  parte de la alucinacion de la ia si se conserva se debe cambiar 
    #Joe:  para checar en los dos, no es necesario
    if catalogo['SKUMERCK'].duplicated().any():
        repetidos = catalogo.loc[catalogo['SKUMERCK'].duplicated(), 'SKUMERCK'].tolist()
        raise ValueError("El catalogo tiene SKU repetidos y la demanda se "
                         f"contaria dos veces: {repetidos}")
    #Joe: busca valores ausentes en el forecast, no es necesario en este momento

    if 'SKUMERCK' not in fcst.columns:
        raise ValueError(f"El forecast no tiene la columna '{'SKUMERCK'}'.")

    log(f"Forecast: {len(fcst)} filas | Catalogo: {len(catalogo)} filas")
    return fcst, catalogo


# ==============================================================================
# 2. Ventana de meses
# ==============================================================================

def columnas_de_mes(columnas):
    """Devuelve los encabezados de mes que son demanda futura.

    Se recorren los encabezados de derecha a izquierda y se guardan los que
    pandas pueda leer como fecha, hasta topar con el primero que ya no es
    futuro. Como el forecast trae dos bloques de meses (cajas y piezas), este
    recorrido inverso cae en el segundo bloque, que es el de piezas.
    """
    ''''''
    #Joe: se agrego el codigo viejo como comentario y se dejo el de la ia
    #Joe: el codigo de la ia es practicamente lo mismo pero menos lineas
    #Joe: ya se verifico que no afecta el funcionamiento y salen los meses correctos. 
    '''
        #Calculo de mes y año actual
    Month_today=pd.to_datetime("today").month
    Year_today=pd.to_datetime("today").year
    demand_today=pd.to_datetime(str(Year_today) + "-01-" + str(Month_today), format="%Y-%m-%d")
    demand_today=str(demand_today)[0:10]

    #seleccion de meses con demanda futura en el FCST
    matching_cols = []
    for i in range(len(FCST.columns) - 1, -1, -1):  # from [-1] backwards
        col = FCST.columns[i]
        parsed = pd.to_datetime(col, errors='coerce')
        if pd.notna(parsed):
            matching_cols.append(col)
            if parsed.month <= Month_today and parsed.year <= Year_today:
                break  #
    matching_cols = list(reversed(matching_cols))
    '''
    # codigo nuevo de la ia
    # Joe: se estaba importando la variable hoy vacia, que se estaba calculando en vano 
    # Joe: la eliminamos y la conservamos para que se calcule localmente en la función y ya 
    # Joe: version de la ia: hoy = pd.Timestamp.today() if hoy is None else hoy
    # Joe: se adjuntaron dos secciones de las funciones que usan esta sub función.
    hoy = pd.Timestamp.today()
    seleccion = []

    for columna in reversed(list(columnas)):
        fecha = pd.to_datetime(columna, errors="coerce")
        if pd.isna(fecha):
            continue
        seleccion.append(columna)
        if (fecha.year, fecha.month) <= (hoy.year, hoy.month):
            break
    
    seleccion.reverse()

    fechas = pd.to_datetime(pd.Series(seleccion), errors="coerce")
    if not seleccion or fechas.dt.year.min() < 2000:
        raise ValueError("No se detectaron encabezados de mes validos en el "
                         "forecast. Revise que los encabezados de mes sean "
                         "fechas de Excel.")
    
    meses = [f.strftime(FORMATO_MES) for f in fechas]
    #seleccion.to_csv("seleccion.csv", index=False)
    
    return seleccion, meses


def listar_meses(ruta_fcst, hoja_fcst):
    """Etiquetas de mes que ofrece la hoja, leyendo solo los encabezados.
    Sirve para llenar el combo de la app sin cargar todo el libro. Si la hoja
    no trae fechas validas regresa una lista vacia.
    """
    #joe: establece una funcion llamando a la funcion previa para extraer los meses
    # despues los formatea a formato mes y pone como minimo que sean mayores al año 2000
    # se usa para las opciones de mes a analizar.
    #Joe: ya se cambio el codigo excesivo a la funcino de columnas_de_mes
    encabezados = pd.read_excel(ruta_fcst, sheet_name=hoja_fcst, nrows=0).columns
    try:
        _, meses = columnas_de_mes(encabezados)
    except ValueError:
        return []
    return meses

def preparar_demanda(fcst):
    """Recorta el forecast a SKU + meses futuros y renombra los encabezados.
    Regresa (demanda, meses), donde `meses` son etiquetas tipo '08-2026'.
    """
    #JOE: llama la funcion de columnas de mes
    columnas, meses = columnas_de_mes(fcst.columns)
    '''
    ESTE CODIGO SE ENCUENTRA EN DOS FUNCIONES DIFERENTES QUE PROCESAN LOS MISMOS DATOS QUE 
    COLUMNAS DE MES, SE VA A INTEGRAR A COLUMNAS DE MES CON UN MENSAJE GENERALIZADO. 

    if not columnas:
        raise ValueError("No se detectaron columnas de fecha en el forecast. "
                         "Revise que los encabezados de mes sean fechas de Excel.")

    # se paso esta parte a la funcino de columnas ya que tiene mas sentido que la 
    # ejecute en la limpieza 
    # y no como un paso perdido
    fechas = pd.to_datetime(pd.Series(columnas), errors="coerce")
    if fechas.dt.year.min() < 2000:
        raise ValueError("Los encabezados de mes no se leyeron como fechas "
                         "validas. Revise la fila de encabezados del forecast.")
'''
    demanda = fcst[['SKUMERCK'] + columnas].dropna(subset=['SKUMERCK']).copy()
    demanda.columns = ['SKUMERCK'] + meses
    demanda[meses] = demanda[meses].astype(float)
    log(f"Ventana de demanda: {len(meses)} meses ({meses[0]} a {meses[-1]})")
    
    return demanda, meses

#la seccion dos si sirve y esta haciendo correctamente su trabajo, el codigo 
#de hecho encontro un bug con el original que no estaba contemplado. 
# Todavia se podria retrabajar y perfeccionar pero por ahora se juega. 

# ==============================================================================
# 3. Cruce con el catalogo y consolidacion
# ==============================================================================

def detalle_por_sku(catalogo, demanda, meses, fcst=None):
    """Una fila por SKU con su familia, su linea y su demanda en piezas.
    Esta tabla es la base de todo: el consolidado sale de agruparla, asi que
    las dos cuadran por construccion.
    """
    tabla = catalogo[['SKUMERCK', "Nombre granel", "Linea Granel 1", "Linea Granel 2"]].merge(
        demanda, on='SKUMERCK', how="left")
    '''
    la siguiente parte es inecesaria y no afecta el codigo en el futuro, solo se conservara
    la seccion de fillna para que no haya nulos en la tabla final.
    sin_forecast = int(tabla[meses[0]].isna().sum())
    if sin_forecast:
        log(f"Aviso: {sin_forecast} SKU del catalogo no aparecen en el forecast "
            "y entran con demanda cero.")
    '''
    tabla[meses] = tabla[meses].astype(float).fillna(0.0)

    # Las dos banderas del catalogo se colapsan en una sola etiqueta de linea.
    #JOE: esta parte me encanto, asi se queda. 

    corre_en_dc2 = (tabla["Linea Granel 1"] == 1) & (tabla["Linea Granel 2"] == 1)
    tabla["DC"] = corre_en_dc2.map({True: "DC2", False: "DC1"})
    tabla = tabla.drop(columns=["Linea Granel 1", "Linea Granel 2"])

    # Datos descriptivos del forecast, solo para que el reporte se lea mejor.
    descriptivas = []
    if fcst is not None:
        descriptivas = [c for c in ("Description", "Units") if c in fcst.columns]
    if descriptivas:
        catalogo_desc = fcst[['SKUMERCK'] + descriptivas].drop_duplicates(subset=['SKUMERCK'])
        tabla = tabla.merge(catalogo_desc, on='SKUMERCK', how="left")
    # la tabla total horizonte practicamente seria el balanceo por producto ya 
    # echa, solo necesita implenetar ese codigo de balanceo para ser veridica. 

    tabla = tabla[["Nombre granel", "DC", 'SKUMERCK'] + descriptivas + meses]

    tabla = tabla.sort_values(["Nombre granel", "DC"],
                              ascending=[True, True])
    log(f"Detalle por SKU: {len(tabla)} renglones")
    
    tabla.to_csv("tabla.csv", index=False)
    return tabla.reset_index(drop=True)


def consolidar(detalle, meses):
    """Suma el detalle por familia de granel y linea."""
    consolidado = detalle.groupby(["Nombre granel", "DC"], as_index=False)[meses].sum()
    log(f"Consolidado: {consolidado["Nombre granel"].nunique()} familias, "
        f"{(consolidado["DC"] == 'DC2').sum()} bloques DC2")
    consolidado.to_csv("consolidado.csv", index=False)
    return consolidado


# ==============================================================================
# 4. Balanceo DC2 -> DC1 y conversion a lotes
# ==============================================================================

def balancear(consolidado, meses, lote=TAMANO_LOTE):
    """Cierra los lotes de DC2 pasando su remanente a DC1 y convierte a lotes.
    Por cada familia y cada mes:
        residuo = demanda_DC2 % lote
        DC2 se queda con lotes enteros y DC1 absorbe el residuo.
    Regresa (lotes, residuos). `residuos` esta en piezas y dice cuanto falta
    para cerrar el ultimo lote de cada fila.
    """
    tabla = consolidado.copy()
    #JOE: en ves de utilizar mask usa index y garantiza que no se repita nada, sirve igual
    for familia, grupo in tabla.groupby("Nombre granel"):
        filas_dc2 = grupo.index[grupo["DC"] == "DC2"]
        filas_dc1 = grupo.index[grupo["DC"] == "DC1"]
        if len(filas_dc2) > 1 or len(filas_dc1) > 1:
            raise ValueError(f"La familia '{familia}' tiene mas de una fila por "
                         "linea en el consolidado; revise el catalogo.")
        if len(filas_dc2) == 0:
            continue
        if len(filas_dc1) == 0:
            # Sin bloque DC1 no hay donde depositar el remanente: se deja
            # intacto para no perder demanda.
            log(f"Aviso: la familia '{familia}' tiene DC2 pero no DC1, "
                "no se transfiere su remanente.")
            continue

        origen, destino = filas_dc2[0], filas_dc1[0]
        for mes in meses:
            residuo = tabla.at[origen, mes] % lote
            tabla.at[origen, mes] -= residuo
            tabla.at[destino, mes] += residuo

    # Faltante en piezas para cerrar el ultimo lote, antes de dividir.
    residuos = tabla.copy()
    excedente = tabla[meses] % lote
    #despreciable = (faltante <= lote * tolerancia) | (faltante >= lote * (1 - tolerancia))
    residuos[meses] = excedente
    tabla[meses] = (tabla[meses] / lote)
    return tabla, residuos

# ==============================================================================
# 5. Exportacion a Excel
# ==============================================================================

def filtrar_mes(tabla, mes, meses, quitar=()):
    """Deja solo las columnas descriptivas y la del mes indicado."""
    descriptivas = [c for c in tabla.columns if c not in meses and c not in quitar]
    return tabla[descriptivas + [mes]].copy()


def armar_resumen(meses, mes, detalle, lotes, lote, origen):
    """Tabla de trazabilidad que encabeza el libro."""
    filas = [
        ("Fecha de ejecucion", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Jeringas por lote", f"{lote:,}"),
        ("Meses procesados", f"{len(meses)} ({meses[0]} a {meses[-1]})"),
        ("Familias de granel", str(lotes["Nombre granel"].nunique())),
        ("SKU considerados", str(len(detalle))),
        ("Archivo de forecast", origen["fcst"]),
        ("Archivo de catalogo", origen["catalogo"])]
    return pd.DataFrame(filas, columns=["Concepto", "Valor"])

def dar_formato(hoja, tabla, formato_numeros):
    """Encabezado verde, anchos de columna, filtro y formato numerico."""
    for celda in hoja[1]:
        celda.font = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=VERDE_MERCK)
        celda.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
    hoja.row_dimensions[1].height = 28
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.font = Font(name=FUENTE, size=10)
            if isinstance(celda.value, (int, float)):
                celda.number_format = formato_numeros
    for i, columna in enumerate(tabla.columns, start=1):
        ancho_datos = int(tabla[columna].astype(str).str.len().max()) if len(tabla) else 0
        ancho = min(max(len(str(columna)) + 4, ancho_datos + 2, 10), 42)
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.freeze_panes = "C2"
    hoja.auto_filter.ref = hoja.dimensions

def exportar_excel(ruta_salida, resumen, lotes, detalle, residuos):
    """Escribe el libro con las cuatro hojas del proceso."""
    ruta_salida = Path(ruta_salida).with_suffix(".xlsx")
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    hojas = {
        "Resumen": (resumen, "@"),
        "Lotes por familia": (lotes, FORMATO_LOTES),
        "Detalle SKU": (detalle, FORMATO_PIEZAS),
        "Residuos": (residuos, FORMATO_PIEZAS),}
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        for nombre, (tabla, formato) in hojas.items():
            tabla.to_excel(writer, sheet_name=nombre, index=False)
            dar_formato(writer.book[nombre], tabla, formato)
        writer.book["Resumen"].column_dimensions["B"].width = 60
    log(f"Excel generado: {ruta_salida}")
    return ruta_salida

# ==============================================================================
# 6. Proceso completo
# ==============================================================================
def ejecutar(ruta_fcst, hoja_fcst, ruta_catalogo, hoja_catalogo, ruta_salida,
             lote=TAMANO_LOTE, mes=None):
    """Corre todo el proceso y escribe el Excel.

    `mes` limita las hojas 'Detalle SKU' y 'Residuos' a ese mes; los lotes
    siempre se calculan sobre todo el horizonte.

    Regresa un diccionario con las tablas, los meses y los totales por linea.
    """
    log(f"Inicio | lote={lote:,} jeringas")

    fcst, catalogo = leer_entradas(ruta_fcst, hoja_fcst, ruta_catalogo, hoja_catalogo)
    demanda, meses = preparar_demanda(fcst)

    if mes and mes not in meses:
        raise ValueError(f"El mes '{mes}' no esta en el horizonte "
                         f"({meses[0]} a {meses[-1]}).")

    detalle = detalle_por_sku(catalogo, demanda, meses, fcst)
    consolidado = consolidar(detalle, meses)
    lotes, residuos = balancear(consolidado, meses, lote)

    detalle_salida, residuos_salida = detalle, residuos
    if mes:
        detalle_salida = filtrar_mes(detalle, mes, meses)
        detalle_salida = detalle_salida.sort_values(
            ["Nombre granel", "DC", mes], ascending=[True, True, False])
        residuos_salida = filtrar_mes(residuos, mes, meses)
        log(f"Detalle y residuos recortados al mes {mes}")

    origen = {
        "fcst": f"{Path(ruta_fcst).name} (hoja '{hoja_fcst}')",
        "catalogo": f"{Path(ruta_catalogo).name} (hoja '{hoja_catalogo}')",
    }
    resumen = armar_resumen(meses, mes, detalle, lotes, lote, origen)
    exportar_excel(ruta_salida, resumen, lotes, detalle_salida, residuos_salida)

    total_dc1 = float(lotes.loc[lotes["DC"] == "DC1", meses].to_numpy().sum())
    total_dc2 = float(lotes.loc[lotes["DC"] == "DC2", meses].to_numpy().sum())
    log(f"Fin | {total_dc1:,.2f} lotes en DC1 | {total_dc2:,.2f} lotes en DC2")

    return {
        "lotes": lotes,
        "detalle": detalle,
        "residuos": residuos,
        "meses": meses,}
